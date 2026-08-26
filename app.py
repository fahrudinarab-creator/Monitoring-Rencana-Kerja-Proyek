"""
RKP Monitor — Dashboard Monitoring Rencana Kerja Proyek
Jalankan lokal:  streamlit run app.py
Deploy: push ke GitHub lalu hubungkan repo di https://share.streamlit.io
"""

import re
import io
import hashlib
from pathlib import Path
from datetime import datetime

import openpyxl
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st

# Folder di repo GitHub tempat file .xlsx RKP disimpan.
# Update dashboard = tambah/ganti file .xlsx di folder ini lalu commit ke GitHub.
DATA_DIR = Path(__file__).parent / "data"

# ============================================================
# KONFIGURASI HALAMAN & TEMA
# ============================================================
st.set_page_config(
    page_title="RKP Monitor",
    page_icon="🌴",
    layout="wide",
    initial_sidebar_state="expanded",
)

FOREST = "#1F4D36"
FOREST_LIGHT = "#3C7A5A"
GOLD = "#C08A2E"
RUST = "#B25330"
PALETTE = ["#1F4D36", "#C08A2E", "#3C7A5A", "#B25330", "#7C9A85", "#8C6A2E", "#4E6B57", "#D9AE63"]

st.markdown(
    f"""
    <style>
    /* Paksa tema terang ini terlepas dari mode gelap/terang browser/OS, supaya
       tidak bergantung pada .streamlit/config.toml ikut ter-upload atau tidak. */

    html, body, .stApp, [data-testid="stAppViewContainer"], [data-testid="stHeader"],
    [data-testid="stMain"], [data-testid="stBottomBlockContainer"] {{
        background-color: #F4F5EF !important;
        color: #1B2A1E !important;
    }}

    h1, h2, h3, h4, h5, h6,
    [data-testid="stMarkdownContainer"] h1, [data-testid="stMarkdownContainer"] h2,
    [data-testid="stMarkdownContainer"] h3, [data-testid="stMarkdownContainer"] h4 {{
        font-family: 'Georgia', serif !important; color: {FOREST} !important;
    }}

    [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] li,
    [data-testid="stMarkdownContainer"] span,
    .stCaption, [data-testid="stCaptionContainer"],
    label, .stRadio label, .stRadio span {{
        color: #1B2A1E !important;
    }}
    [data-testid="stCaptionContainer"] p {{ color: #6C7566 !important; }}

    div[data-testid="stMetric"] {{
        background: #FFFFFF !important; border: 1px solid #E1E3D9; border-radius: 12px;
        padding: 14px 16px; box-shadow: 0 1px 2px rgba(27,42,30,0.04);
    }}
    div[data-testid="stMetricLabel"] p {{ color: #6C7566 !important; font-size: 12.5px; }}
    div[data-testid="stMetricValue"] {{ color: {FOREST} !important; }}
    div[data-testid="stMetricDelta"] {{ color: {FOREST_LIGHT} !important; }}

    /* Tabs */
    button[data-baseweb="tab"] p {{ color: #6C7566 !important; font-weight: 600; }}
    button[data-baseweb="tab"][aria-selected="true"] p {{ color: {FOREST} !important; }}
    [data-testid="stTabs"] {{ background: transparent !important; }}
    div[data-baseweb="tab-highlight"] {{ background-color: {GOLD} !important; }}
    div[data-baseweb="tab-border"] {{ background-color: #E1E3D9 !important; }}

    /* Sidebar — bertema hijau tua supaya konsisten sebagai identitas brand */
    section[data-testid="stSidebar"] {{
        background-color: #17372A !important;
    }}
    section[data-testid="stSidebar"] * {{
        color: #F4F5EF !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p {{
        color: #B9D6C4 !important;
    }}
    section[data-testid="stSidebar"] .stButton button {{
        background-color: {GOLD} !important; color: #2A1D06 !important; border: none !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stExpander"] {{
        border: 1px solid rgba(255,255,255,0.2) !important; border-radius: 10px;
    }}
    section[data-testid="stSidebar"] [data-testid="stAlertContainer"] {{
        color: #1B2A1E !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stAlertContainer"] * {{
        color: #1B2A1E !important;
    }}

    /* Dataframe / table */
    [data-testid="stDataFrame"] {{ color: #1B2A1E !important; }}

    .badge-wait {{ background:#F5E4DA; color:{RUST}; padding:3px 10px; border-radius:99px; font-size:11.5px; font-weight:700; }}
    .badge-ok {{ background:#E4EEE7; color:{FOREST}; padding:3px 10px; border-radius:99px; font-size:11.5px; font-weight:700; }}
    .footnote {{ background:#E4EEE7; border-radius:10px; padding:12px 14px; font-size:13px; color:#3d4a40 !important; }}
    .footnote * {{ color:#3d4a40 !important; }}
    </style>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# HELPER FORMAT
# ============================================================
def fmt_rp(n):
    if n is None or pd.isna(n):
        return "—"
    if abs(n) >= 1e9:
        return f"Rp {n/1e9:,.2f} M".replace(",", "X").replace(".", ",").replace("X", ".")
    if abs(n) >= 1e6:
        return f"Rp {n/1e6:,.1f} Jt".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"Rp {n:,.0f}".replace(",", ".")


def fmt_rp_full(n):
    if n is None or pd.isna(n):
        return "—"
    return f"Rp {n:,.0f}".replace(",", ".")


def fmt_ha(n):
    if n is None or pd.isna(n):
        return "—"
    return f"{n:,.2f} Ha".replace(",", "X").replace(".", ",").replace("X", ".")


# ============================================================
# PARSER — mengikuti struktur sheet "RKP" (lihat dokumentasi di README)
# ============================================================
def norm(v):
    return "" if v is None else str(v).strip().lower()


def find_header_row(rows, limit=25):
    for r in range(min(len(rows), limit)):
        row = rows[r]
        for c, v in enumerate(row):
            if norm(v) == "pekerjaan":
                return r
    return -1


def parse_rkp_rows(rows):
    """rows: list-of-list, 0-indexed, hasil openpyxl iter_rows(values_only=True)."""
    hr = find_header_row(rows)
    if hr == -1:
        return None

    tahun_row = rows[hr] if hr < len(rows) else []
    periode_row = rows[hr + 1] if hr + 1 < len(rows) else []
    sub_row = rows[hr + 2] if hr + 2 < len(rows) else []

    vol_col = next((c for c, v in enumerate(sub_row) if norm(v) == "volume"), -1)
    if vol_col == -1:
        return None
    biaya_col = vol_col + 1

    fisik_cols = [c for c, v in enumerate(sub_row) if norm(v) == "fisik"]
    if not fisik_cols:
        return None

    maxc = max(len(tahun_row), len(periode_row), len(sub_row))
    last_t, last_p = None, None
    filled_t, filled_p = {}, {}
    for c in range(fisik_cols[0], maxc):
        if c < len(tahun_row) and tahun_row[c] not in (None, ""):
            last_t = tahun_row[c]
        if c < len(periode_row) and periode_row[c] not in (None, ""):
            last_p = periode_row[c]
        filled_t[c] = last_t
        filled_p[c] = last_p

    periods = []
    for fc in fisik_cols:
        bc = fc + 1
        tl = str(filled_t.get(fc) or "")
        pl = str(filled_p.get(fc) or "")
        is_cw = bool(re.search(r"catur\s*wulan", pl, re.I))
        year_m = re.search(r"(\d{4})", tl)
        cw_m = re.search(r"catur\s*wulan\s*(\d+)", pl, re.I)
        year = int(year_m.group(1)) if year_m else None
        cw = int(cw_m.group(1)) if cw_m else None
        display = f"{year or '?'} CW{cw or '?'}" if is_cw else pl.strip()
        periods.append(
            dict(fisik_col=fc, biaya_col=bc, tahun=tl.strip(), periode=pl.strip(),
                 is_cw=is_cw, year=year, cw=cw, sort_key=(year or 0) * 10 + (cw or 0), display=display)
        )
    cw_periods = [p for p in periods if p["is_cw"]]

    grand_row = -1
    for r in range(hr + 3, len(rows)):
        row = rows[r]
        label = row[1] if len(row) > 1 else None
        label2 = row[2] if len(row) > 2 else None
        if (label and re.search(r"grand.?total", str(label), re.I)) or (
            label2 and re.search(r"grand.?total", str(label2), re.I)
        ):
            grand_row = r
            break

    end = grand_row if grand_row != -1 else len(rows)
    items = []
    for r in range(hr + 3, end):
        row = rows[r]
        name = row[2] if len(row) > 2 else None
        vol = row[vol_col] if vol_col < len(row) else None
        biaya = row[biaya_col] if biaya_col < len(row) else None
        if name is None or str(name).strip() == "":
            continue
        if not isinstance(vol, (int, float)) and not isinstance(biaya, (int, float)):
            continue
        per = []
        for p in cw_periods:
            f = row[p["fisik_col"]] if p["fisik_col"] < len(row) else None
            b = row[p["biaya_col"]] if p["biaya_col"] < len(row) else None
            per.append(dict(key=p["display"], sort_key=p["sort_key"],
                             fisik=f if isinstance(f, (int, float)) else 0,
                             biaya=b if isinstance(b, (int, float)) else 0))
        items.append(dict(
            no=row[1] if len(row) > 1 else None,
            nama=str(name).strip(),
            volume_ha=vol if isinstance(vol, (int, float)) else None,
            biaya_rencana=biaya if isinstance(biaya, (int, float)) else None,
            rp_per_ha=(biaya / vol) if isinstance(biaya, (int, float)) and isinstance(vol, (int, float)) and vol else None,
            periods=per,
        ))

    grand = None
    if grand_row != -1:
        row = rows[grand_row]
        gvol = row[vol_col] if vol_col < len(row) else None
        gbiaya = row[biaya_col] if biaya_col < len(row) else None
        gper = []
        for p in cw_periods:
            f = row[p["fisik_col"]] if p["fisik_col"] < len(row) else None
            b = row[p["biaya_col"]] if p["biaya_col"] < len(row) else None
            gper.append(dict(key=p["display"], sort_key=p["sort_key"],
                              fisik=f if isinstance(f, (int, float)) else 0,
                              biaya=b if isinstance(b, (int, float)) else 0))
        grand = dict(volume_ha=gvol if isinstance(gvol, (int, float)) else None,
                     biaya_rencana=gbiaya if isinstance(gbiaya, (int, float)) else None,
                     periods=gper)

    return dict(items=items, grand=grand)


def extract_meta(rows, file_name):
    company, desc, luas_text, periode_text = None, None, None, None
    for r in range(min(len(rows), 10)):
        row = rows[r]
        for c, v in enumerate(row):
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            if re.search(r"rencana kerja proyek", s, re.I):
                continue
            if re.match(r"^pt[\s.]", s, re.I) and not company:
                company = s
                continue
            if re.search(r"luas", s, re.I):
                for c2 in range(c + 1, len(row)):
                    if row[c2] not in (None, "") and str(row[c2]).strip() != "":
                        luas_text = str(row[c2]).strip()
                        break
            if re.search(r"tahun|periode", s, re.I) and re.search(r"\d{4}", s) and not periode_text:
                periode_text = s
            if 2 <= r <= 4 and not desc and not re.search(r"luas|tahun", s, re.I) and not re.match(r"^pt[\s.]", s, re.I):
                desc = s

    luas_num = None
    if luas_text:
        m = re.search(r"[\d.]+", luas_text.replace(".", "").replace(",", "."))
        if m:
            try:
                luas_num = float(m.group(0))
            except ValueError:
                luas_num = None

    name = re.sub(r"\.xlsx$", "", file_name, flags=re.I)
    name = re.sub(r"^RKP[_\s]?", "", name, flags=re.I)
    name = re.sub(r"_+", " ", name).strip()

    return dict(name=name, company=company or "—", desc=desc or "", luas_num=luas_num,
                luas_text=luas_text, periode_text=periode_text)


def try_parse_realisasi(wb):
    real_sheet = next((n for n in wb.sheetnames if re.search(r"realisasi|aktual", n, re.I)), None)
    if not real_sheet:
        return dict(status="none")
    try:
        ws = wb[real_sheet]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        parsed = parse_rkp_rows(rows)
        if not parsed:
            return dict(status="unrecognized", sheet=real_sheet)
        return dict(status="ok", data=parsed, sheet=real_sheet)
    except Exception:
        return dict(status="unrecognized", sheet=real_sheet)


def parse_workbook(file_bytes, file_name):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "RKP" not in wb.sheetnames:
        return dict(error='Sheet "RKP" tidak ditemukan di file ini.')
    ws = wb["RKP"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    rencana = parse_rkp_rows(rows)
    if not rencana:
        return dict(error='Format sheet "RKP" tidak dikenali (header "Pekerjaan"/"Volume"/"Fisik" tidak ditemukan).')
    meta = extract_meta(rows, file_name)
    realisasi = try_parse_realisasi(wb)
    return dict(id=file_name, file_name=file_name, updated_at=datetime.now().isoformat(),
                meta=meta, rencana=rencana, realisasi=realisasi)


# ============================================================
# DERIVED HELPERS
# ============================================================
def total_rencana(p):
    return p["rencana"]["grand"]["biaya_rencana"] if p["rencana"]["grand"] else None


def luas_proj(p):
    g = p["rencana"]["grand"]
    return g["volume_ha"] if g and g["volume_ha"] else p["meta"]["luas_num"]


def rp_per_ha(p):
    t, l = total_rencana(p), luas_proj(p)
    return t / l if t and l else None


def has_realisasi(p):
    return p["realisasi"]["status"] == "ok"


def all_period_keys(projects):
    seen = {}
    for p in projects.values():
        g = p["rencana"]["grand"]
        if not g:
            continue
        for pd_ in g["periods"]:
            if pd_["key"] not in seen or pd_["sort_key"] < seen[pd_["key"]]:
                seen[pd_["key"]] = pd_["sort_key"]
    return [k for k, _ in sorted(seen.items(), key=lambda x: x[1])]


# ============================================================
# DATA DARI REPO GITHUB (folder data/)
# ============================================================
def _repo_cache_key():
    """Kunci cache berbasis nama+ukuran+waktu-modifikasi file di folder data/.
    Berubah otomatis begitu file di-update lewat GitHub -> cache re-parse."""
    if not DATA_DIR.exists():
        return "no-data-dir"
    parts = []
    for fp in sorted(DATA_DIR.glob("*.xlsx")):
        stat = fp.stat()
        parts.append(f"{fp.name}:{stat.st_mtime}:{stat.st_size}")
    return hashlib.md5("|".join(parts).encode()).hexdigest()


@st.cache_data(show_spinner="Membaca file RKP dari repo GitHub...")
def load_repo_projects(_cache_key):
    result = {}
    errs = []
    if not DATA_DIR.exists():
        return result, errs
    for fp in sorted(DATA_DIR.glob("*.xlsx")):
        try:
            parsed = parse_workbook(fp.read_bytes(), fp.name)
            if "error" in parsed:
                errs.append(f"{fp.name}: {parsed['error']}")
                continue
            result[parsed["id"]] = parsed
        except Exception as e:
            errs.append(f"{fp.name}: gagal dibaca ({e})")
    return result, errs


# ============================================================
# STATE
# ============================================================
if "session_projects" not in st.session_state:
    st.session_state.session_projects = {}  # file uji coba sementara (tidak permanen)

repo_projects, repo_errs = load_repo_projects(_repo_cache_key())

# ============================================================
# SIDEBAR
# ============================================================
with st.sidebar:
    st.markdown("### 🌴 RKP Monitor")

    if repo_projects:
        st.success(f"📁 {len(repo_projects)} file RKP dimuat dari repo GitHub (folder `data/`)")
    elif DATA_DIR.exists():
        st.warning("Folder `data/` ada tapi belum berisi file .xlsx.")
    else:
        st.info("Folder `data/` belum ada di repo. Lihat README untuk cara menambahkannya.")
    for e in repo_errs:
        st.warning(e)

    if st.button("🔄 Muat ulang data dari GitHub", use_container_width=True):
        load_repo_projects.clear()
        st.rerun()

    st.divider()
    st.caption("Untuk update data secara permanen: ganti/tambah file `.xlsx` di folder `data/` pada repo GitHub, lalu commit — dashboard otomatis membaca versi terbaru dalam beberapa menit.")

    with st.expander("🧪 Uji coba file lain (sementara, tidak permanen)"):
        uploaded = st.file_uploader("Upload file RKP", type=["xlsx"], accept_multiple_files=True, label_visibility="collapsed")
        if uploaded:
            ok, errs = 0, []
            for f in uploaded:
                try:
                    result = parse_workbook(f.read(), f.name)
                    if "error" in result:
                        errs.append(f"{f.name}: {result['error']}")
                        continue
                    st.session_state.session_projects[result["id"]] = result
                    ok += 1
                except Exception as e:
                    errs.append(f"{f.name}: gagal dibaca ({e})")
            if ok:
                st.success(f"{ok} file berhasil diproses (sesi ini saja).")
            for e in errs:
                st.warning(e)
        if st.session_state.session_projects:
            st.caption(f"{len(st.session_state.session_projects)} file uji coba aktif")
            if st.button("Hapus file uji coba", use_container_width=True):
                st.session_state.session_projects = {}
                st.rerun()

# Gabungkan: data dari repo GitHub (utama) + file uji coba sesi (opsional, menimpa nama file yang sama)
projects_all = {**repo_projects, **st.session_state.session_projects}

# ============================================================
# SIDEBAR — FILTER
# ============================================================
projects = projects_all
if projects_all:
    with st.sidebar:
        st.divider()
        st.markdown("#### 🔎 Filter")

        companies = sorted({p["meta"]["company"] for p in projects_all.values()})
        selected_companies = st.multiselect(
            "Perusahaan", companies, default=companies, key="filter_company"
        )

        projects_by_company = {
            k: v for k, v in projects_all.items() if v["meta"]["company"] in selected_companies
        }

        project_names = sorted({p["meta"]["name"] for p in projects_by_company.values()})
        selected_projects = st.multiselect(
            "Proyek", project_names, default=project_names, key="filter_project"
        )

        projects = {
            k: v for k, v in projects_by_company.items() if v["meta"]["name"] in selected_projects
        }

        if projects_all and not projects:
            st.warning("Tidak ada proyek yang cocok dengan filter ini.")

        if projects:
            st.markdown("#### 👁️ Tampilan")
            view_options = ["📊 Ringkasan"] + [p["meta"]["name"] for p in projects.values()]
            if st.session_state.get("view_select") not in view_options:
                st.session_state["view_select"] = view_options[0]
            st.selectbox("Pilih tampilan", view_options, key="view_select", label_visibility="collapsed")


# ============================================================
# MAIN
# ============================================================
st.title("🌴 Monitoring RKP")
st.caption("Dashboard konsolidasi Rencana Kerja Proyek — banding biaya & capaian fisik lintas proyek.")

if not projects_all:
    st.info(
        "⬅️ Belum ada data. Tambahkan file `.xlsx` RKP ke folder **`data/`** di repo GitHub lalu "
        "commit, atau upload file uji coba lewat panel kiri untuk mulai memantau."
    )
    st.stop()

if not projects:
    st.warning("Tidak ada proyek yang cocok dengan filter Perusahaan/Proyek yang dipilih di sidebar. Coba longgarkan filternya.")
    st.stop()

selected_view = st.session_state.get("view_select", "📊 Ringkasan")

# ================================================================
# RINGKASAN (semua proyek terfilter)
# ================================================================
if selected_view == "📊 Ringkasan":
    total_biaya = sum((total_rencana(p) or 0) for p in projects.values())
    total_luas = sum((luas_proj(p) or 0) for p in projects.values())
    avg_rp_ha = total_biaya / total_luas if total_luas else None
    n_real = sum(1 for p in projects.values() if has_realisasi(p))

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Jumlah Proyek", len(projects), f"{n_real} dengan realisasi")
    c2.metric("Total Luas", fmt_ha(total_luas))
    c3.metric("Total Biaya Rencana", fmt_rp(total_biaya), fmt_rp_full(total_biaya))
    c4.metric("Rata-rata Biaya / Ha", fmt_rp(avg_rp_ha))

    st.markdown("#### Perbandingan Biaya Antar Proyek")
    mode = st.radio("Mode", ["Total Biaya", "Biaya / Ha"], horizontal=True, label_visibility="collapsed")
    names = [p["meta"]["name"] for p in projects.values()]
    vals = [
        (total_rencana(p) or 0) if mode == "Total Biaya" else (rp_per_ha(p) or 0)
        for p in projects.values()
    ]
    fig = go.Figure(go.Bar(
        x=vals, y=names, orientation="h",
        marker_color=[PALETTE[i % len(PALETTE)] for i in range(len(names))],
        text=[fmt_rp(v) for v in vals], textposition="outside",
    ))
    fig.update_layout(height=100 + 60 * len(names), margin=dict(l=10, r=40, t=10, b=10),
                       xaxis_title="Rp", plot_bgcolor="white", paper_bgcolor="white",
                       font=dict(color="#1B2A1E", size=13),
                       xaxis=dict(gridcolor="#EEF0E8", color="#1B2A1E"),
                       yaxis=dict(color="#1B2A1E", automargin=True))
    st.plotly_chart(fig, use_container_width=True, theme=None)

    st.markdown("#### Target Fisik per Periode (Catur Wulan) — Semua Proyek")
    keys = all_period_keys(projects)
    fig2 = go.Figure()
    for i, p in enumerate(projects.values()):
        g = p["rencana"]["grand"]
        m = {pd_["key"]: pd_["fisik"] for pd_ in (g["periods"] if g else [])}
        fig2.add_bar(name=p["meta"]["name"], x=keys, y=[m.get(k, 0) for k in keys],
                     marker_color=PALETTE[i % len(PALETTE)])
    fig2.update_layout(barmode="stack", height=380, margin=dict(l=10, r=10, t=10, b=10),
                        yaxis_title="Ha", plot_bgcolor="white", paper_bgcolor="white",
                        legend=dict(orientation="h", y=-0.3, font=dict(color="#1B2A1E", size=11)),
                        font=dict(color="#1B2A1E", size=13),
                        xaxis=dict(gridcolor="#EEF0E8", color="#1B2A1E"),
                        yaxis=dict(gridcolor="#EEF0E8", color="#1B2A1E", automargin=True))
    st.plotly_chart(fig2, use_container_width=True, theme=None)

    st.markdown("#### Daftar Proyek")
    cols = st.columns(3)
    for i, p in enumerate(projects.values()):
        with cols[i % 3]:
            real = has_realisasi(p)
            badge = '<span class="badge-ok">Ada realisasi</span>' if real else '<span class="badge-wait">Rencana saja</span>'
            st.markdown(
                f"""
                <div style="background:#fff;border:1px solid #E1E3D9;border-radius:12px;padding:14px 16px;margin-bottom:12px;">
                <div style="display:flex;justify-content:space-between;align-items:flex-start;">
                  <div>
                    <b>{p['meta']['name']}</b><br>
                    <span style="color:#6C7566;font-size:12.5px;">{p['meta']['company']}</span>
                  </div>
                  {badge}
                </div>
                <div style="margin-top:8px;font-size:13px;color:#6C7566;">📐 {fmt_ha(luas_proj(p))} &nbsp;·&nbsp; 📅 {p['meta']['periode_text'] or '—'}</div>
                <div style="margin-top:8px;display:flex;justify-content:space-between;">
                  <div><b style="color:{FOREST};">{fmt_rp(total_rencana(p))}</b><br><span style="font-size:11px;color:#6C7566;">total rencana</span></div>
                  <div style="text-align:right;"><b style="color:{FOREST};">{fmt_rp(rp_per_ha(p))}</b><br><span style="font-size:11px;color:#6C7566;">per Ha</span></div>
                </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

# ================================================================
# DETAIL SATU PROYEK (dipilih lewat dropdown sidebar)
# ================================================================
else:
    p = next((pp for pp in projects.values() if pp["meta"]["name"] == selected_view), None)
    if p is None:
        st.warning("Proyek tidak ditemukan pada hasil filter saat ini. Kembali ke Ringkasan.")
        st.stop()

    real = has_realisasi(p)
    total = total_rencana(p)
    luas = luas_proj(p)

    st.subheader(p["meta"]["name"])
    st.caption(f"{p['meta']['company']}" + (f" · {p['meta']['desc']}" if p["meta"]["desc"] else "") +
               (f" · {p['meta']['periode_text']}" if p["meta"]["periode_text"] else ""))

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Luas", fmt_ha(luas))
    c2.metric("Total Biaya Rencana", fmt_rp(total), fmt_rp_full(total))
    c3.metric("Biaya / Ha", fmt_rp(rp_per_ha(p)))
    c4.metric("Status Realisasi", "Tersedia" if real else "Belum ada")

    if not real:
        st.markdown(
            """<div class="footnote">📋 File ini belum berisi data realisasi (kolom/sheet "Realisasi" tidak
            terdeteksi). Tambahkan sheet baru bernama mengandung kata <b>Realisasi</b> dengan struktur tabel
            yang mirip sheet RKP, lalu upload ulang — dashboard otomatis akan menampilkan perbandingan
            Rencana vs Realisasi di sini.</div>""",
            unsafe_allow_html=True,
        )

    st.markdown("#### Biaya per Periode (Catur Wulan)")
    g = p["rencana"]["grand"]
    keys = [pd_["key"] for pd_ in g["periods"]] if g else []
    rencana_vals = [pd_["biaya"] for pd_ in g["periods"]] if g else []
    fig3 = go.Figure()
    fig3.add_bar(name="Rencana", x=keys, y=rencana_vals, marker_color=FOREST)
    if real and p["realisasi"]["data"]["grand"]:
        rmap = {pd_["key"]: pd_["biaya"] for pd_ in p["realisasi"]["data"]["grand"]["periods"]}
        fig3.add_bar(name="Realisasi", x=keys, y=[rmap.get(k, 0) for k in keys], marker_color=GOLD)
    fig3.update_layout(barmode="group", height=320, margin=dict(l=10, r=10, t=10, b=10),
                        plot_bgcolor="white", paper_bgcolor="white",
                        legend=dict(orientation="h", y=-0.2, font=dict(color="#1B2A1E")),
                        font=dict(color="#1B2A1E", size=13),
                        xaxis=dict(gridcolor="#EEF0E8", color="#1B2A1E"),
                        yaxis=dict(gridcolor="#EEF0E8", color="#1B2A1E"))
    st.plotly_chart(fig3, use_container_width=True, key=f"biaya_{p['id']}", theme=None)

    st.markdown("#### Komposisi Biaya per Pekerjaan")
    items = [it for it in p["rencana"]["items"] if it["biaya_rencana"]]
    items = sorted(items, key=lambda x: -x["biaya_rencana"])
    fig4 = px.pie(
        names=[it["nama"] for it in items],
        values=[it["biaya_rencana"] for it in items],
        color_discrete_sequence=PALETTE, hole=0.55,
    )
    fig4.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10),
                        plot_bgcolor="white", paper_bgcolor="white",
                        font=dict(color="#1B2A1E", size=13),
                        legend=dict(font=dict(color="#1B2A1E")))
    fig4.update_traces(textfont=dict(color="#1B2A1E"))
    st.plotly_chart(fig4, use_container_width=True, key=f"comp_{p['id']}", theme=None)

    st.markdown("#### Rincian Pekerjaan")
    real_map = {}
    if real:
        for it in p["realisasi"]["data"]["items"]:
            real_map[it["nama"]] = it["biaya_rencana"]

    rows_table = []
    for it in p["rencana"]["items"]:
        r_biaya = real_map.get(it["nama"])
        capaian = (r_biaya / it["biaya_rencana"] * 100) if (real and r_biaya and it["biaya_rencana"]) else None
        rows_table.append({
            "No": it["no"],
            "Pekerjaan": it["nama"],
            "Volume (Ha)": it["volume_ha"],
            "Biaya Rencana": it["biaya_rencana"],
            "Rp / Ha": it["rp_per_ha"],
            "Realisasi Biaya": r_biaya if real else None,
            "% Capaian": round(capaian, 1) if capaian is not None else None,
        })
    df = pd.DataFrame(rows_table)
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Volume (Ha)": st.column_config.NumberColumn(format="%.2f"),
            "Biaya Rencana": st.column_config.NumberColumn(format="Rp %d"),
            "Rp / Ha": st.column_config.NumberColumn(format="Rp %d"),
            "Realisasi Biaya": st.column_config.NumberColumn(format="Rp %d"),
            "% Capaian": st.column_config.NumberColumn(format="%.1f%%"),
        },
    )
